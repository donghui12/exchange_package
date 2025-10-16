"""
Excel文件生成模块
生成商品导入模板Excel文件
"""
import os
from datetime import datetime
from typing import Dict, List, Any
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("警告: openpyxl 未安装，Excel生成功能将不可用")
    Workbook = None

from src.config import Config

class ExcelGenerator:
    """Excel文件生成器"""
    
    def __init__(self):
        """初始化生成器"""
        if Workbook is None:
            raise ImportError("openpyxl 未安装，无法使用Excel生成功能")
    
    def create_product_template(self, parser, output_dir: str, download_results: Dict = None) -> bool:
        """
        创建商品导入模板Excel文件
        
        Args:
            parser: 数据解析器实例
            output_dir: 输出目录
            download_results: 下载结果统计
            
        Returns:
            bool: 生成是否成功
        """
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "商品信息"
            
            # 获取商品数据
            goods_info = parser.get_goods_basic_info()
            price_info = parser.get_price_info()
            sku_list = parser.get_sku_info()
            
            # 设置表头
            headers = Config.EXCEL_HEADERS
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            
            # 填充商品基础信息（按15列格式）
            row = 2
            
            # 获取基础数据
            goods_name = goods_info.get('goods_name', '')
            goods_id = goods_info.get('goods_id', '')
            
            # 主商品信息行（第一个SKU或主商品）
            first_sku = sku_list[0] if sku_list else {}
            
            # 填充15列数据
            ws.cell(row=row, column=1, value=goods_name)  # * 产品标题
            ws.cell(row=row, column=2, value="CNY")  # 货币类型
            ws.cell(row=row, column=3, value=f"https://mobile.yangkeduo.com/goods.html?goods_id={goods_id}")  # 货源链接
            ws.cell(row=row, column=4, value="拼多多")  # 货源平台
            ws.cell(row=row, column=5, value=goods_id)  # 产品主编号
            ws.cell(row=row, column=6, value="")  # 详情描述（待填写）
            ws.cell(row=row, column=7, value="")  # 货源类目（待填写）
            ws.cell(row=row, column=8, value="")  # 属性（待填写）
            
            # SKU规格处理（应用编码修复）
            if first_sku:
                specs = first_sku.get('specs', [])
                if len(specs) >= 1:
                    spec_value = specs[0].get('spec_value', '')
                    fixed_spec_value = parser._fix_encoding(str(spec_value)) if spec_value else ''
                    ws.cell(row=row, column=9, value=fixed_spec_value)  # SKU规格1
                if len(specs) >= 2:
                    spec_value = specs[1].get('spec_value', '')
                    fixed_spec_value = parser._fix_encoding(str(spec_value)) if spec_value else ''
                    ws.cell(row=row, column=10, value=fixed_spec_value)  # SKU规格2
                
                # 生成平台SKU（应用编码修复）
                spec_values = []
                for spec in specs:
                    spec_value = spec.get('spec_value', '')
                    if spec_value:
                        # 应用编码修复
                        fixed_spec_value = parser._fix_encoding(str(spec_value))
                        spec_values.append(fixed_spec_value)
                platform_sku = "-".join(spec_values) if spec_values else "默认"
                ws.cell(row=row, column=11, value=platform_sku)  # 平台SKU
                
                ws.cell(row=row, column=12, value=first_sku.get('group_price', 0))  # * SKU售价（已转换为元）
            else:
                ws.cell(row=row, column=9, value="")  # SKU规格1
                ws.cell(row=row, column=10, value="")  # SKU规格2
                ws.cell(row=row, column=11, value="默认")  # 平台SKU
                ws.cell(row=row, column=12, value=price_info.get('min_group_price', 0) / 100)  # * SKU售价（price_info中的价格是分）
            
            ws.cell(row=row, column=13, value=999)  # SKU库存（默认值）
            ws.cell(row=row, column=14, value=0.5)  # SKU重量(KG)（默认值）
            ws.cell(row=row, column=15, value="")  # SKU尺寸(CM)（待填写）
            
            row += 1
            
            # 其他SKU信息行
            for sku in sku_list[1:]:  # 跳过第一个SKU
                ws.cell(row=row, column=1, value=goods_name)  # * 产品标题
                ws.cell(row=row, column=2, value="CNY")  # 货币类型
                ws.cell(row=row, column=3, value="")  # 货源链接（空）
                ws.cell(row=row, column=4, value="")  # 货源平台（空）
                ws.cell(row=row, column=5, value="")  # 产品主编号（空）
                ws.cell(row=row, column=6, value="")  # 详情描述（空）
                ws.cell(row=row, column=7, value="")  # 货源类目（空）
                ws.cell(row=row, column=8, value="")  # 属性（空）
                
                # SKU规格处理（应用编码修复）
                specs = sku.get('specs', [])
                if len(specs) >= 1:
                    spec_value = specs[0].get('spec_value', '')
                    fixed_spec_value = parser._fix_encoding(str(spec_value)) if spec_value else ''
                    ws.cell(row=row, column=9, value=fixed_spec_value)  # SKU规格1
                if len(specs) >= 2:
                    spec_value = specs[1].get('spec_value', '')
                    fixed_spec_value = parser._fix_encoding(str(spec_value)) if spec_value else ''
                    ws.cell(row=row, column=10, value=fixed_spec_value)  # SKU规格2
                
                # 生成平台SKU（应用编码修复）
                spec_values = []
                for spec in specs:
                    spec_value = spec.get('spec_value', '')
                    if spec_value:
                        # 应用编码修复
                        fixed_spec_value = parser._fix_encoding(str(spec_value))
                        spec_values.append(fixed_spec_value)
                platform_sku = "-".join(spec_values) if spec_values else "默认"
                ws.cell(row=row, column=11, value=platform_sku)  # 平台SKU
                
                ws.cell(row=row, column=12, value=sku.get('group_price', 0))  # * SKU售价（已转换为元）
                ws.cell(row=row, column=13, value=999)  # SKU库存（默认值）
                ws.cell(row=row, column=14, value=0.5)  # SKU重量(KG)（默认值）
                ws.cell(row=row, column=15, value="")  # SKU尺寸(CM)（待填写）
                
                row += 1
            
            # 设置列宽（15列）
            column_widths = [20, 10, 30, 10, 15, 25, 15, 25, 12, 12, 15, 12, 10, 12, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # 设置数据行样式
            for row_num in range(2, row):
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin")
                    )
                    
                    # 价格列右对齐
                    if col_num in [12]:  # * SKU售价列
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '0.00'
                    # 重量列右对齐
                    elif col_num in [14]:  # SKU重量(KG)列
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = '0.00'
            
            # 保存文件
            file_path = os.path.join(output_dir, Config.FILE_NAMING['excel_filename'])
            wb.save(file_path)
            
            return True
            
        except Exception as e:
            print(f"生成Excel文件失败: {e}")
            return False